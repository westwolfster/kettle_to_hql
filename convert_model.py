#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
将模型详情表（xls/xlsx）转换为标准实体表模板 xlsx。

用法示例：
  python convert_model.py --file 模型详情-库：ap_tenant_user7-表：kj_e_partner_order_d.xls --output ./out
  python convert_model.py --zip  models.zip --output ./out
"""

import argparse
import os
import re
import zipfile
import tempfile
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# ===================== 配置 =====================
# 输出两张表的固定表头
SHEET1_HEADERS = ["主题(必填)", "实体英文名(必填)", "实体中文名(必填)", "描述"]
SHEET2_HEADERS = [
    "表名(必填)", "实体属性英文名(必填)", "实体属性中文名(必填)",
    "实体属性类型(必填)", "实体属性类型长度(整型数字)", "精度(整型数字)",
    "是否主键(必填)", "是否分区(必填)", "非空(必填)", "描述"
]

# 允许的实体属性类型
ALLOWED_TYPES = {"string", "bigint", "double", "decimal", "boolean", "date", "timestamp"}

# 数字转中文（仅处理开头数字）
DIGIT_TO_CN = {
    "0": "零", "1": "一", "2": "二", "3": "三", "4": "四",
    "5": "五", "6": "六", "7": "七", "8": "八", "9": "九"
}

# ===================== 工具函数 =====================
def to_lower(s):
    """全部转为小写"""
    if pd.isna(s) or s is None:
        return ""
    return str(s).strip().lower()

def clean_chinese_name(s: str) -> str:
    """
    处理“实体属性中文名(必填)”：
    1. 删除字符串中间的空格（或替换为 _，这里选择直接删除）
    2. 不能以数字开头 → 开头连续数字转为中文
    3. 删除标点符号和特殊字符
    """
    if pd.isna(s) or s is None:
        return ""
    s = str(s).strip()

    # 1. 删除所有空白字符
    s = re.sub(r"\s+", "", s)

    # 2. 开头连续数字转中文
    m = re.match(r"^(\d+)(.*)$", s)
    if m:
        digits = m.group(1)
        rest = m.group(2)
        cn_digits = "".join(DIGIT_TO_CN.get(d, d) for d in digits)
        s = cn_digits + rest

    # 3. 只保留中文、英文字母、数字、下划线（删除标点与特殊字符）
    s = re.sub(r"[^\u4e00-\u9fffA-Za-z0-9_]", "", s)

    return s

def normalize_type(t: str) -> str:
    """把字段类型规范到允许的集合中，默认 string"""
    if pd.isna(t) or t is None:
        return "string"
    t = str(t).strip().lower()
    # 常见别名映射
    mapping = {
        "varchar": "string", "char": "string", "text": "string",
        "int": "bigint", "integer": "bigint", "long": "bigint",
        "float": "double", "real": "double",
        "number": "decimal", "numeric": "decimal",
        "bool": "boolean", "bit": "boolean",
        "datetime": "timestamp", "time": "timestamp",
    }
    t = mapping.get(t, t)
    if t not in ALLOWED_TYPES:
        t = "string"
    return t

def is_partition(val) -> str:
    """是否分区字段 → 是/否"""
    if pd.isna(val):
        return "否"
    v = str(val).strip()
    if v in ("是", "Y", "y", "1", "true", "True", "yes", "Yes"):
        return "是"
    return "否"

# ===================== 核心转换逻辑 =====================
def read_input_file(file_path: str) -> pd.DataFrame:
    """读取单个输入文件（支持 .xls / .xlsx）"""
    path = Path(file_path)
    if path.suffix.lower() == ".xls":
        # 旧版 excel 用 xlrd
        df = pd.read_excel(file_path, engine="xlrd", dtype=str)
    else:
        df = pd.read_excel(file_path, engine="openpyxl", dtype=str)

    # 统一列名（去掉前后空格）
    df.columns = [str(c).strip() for c in df.columns]

    # 期望的列名映射（兼容可能的细微差异）
    col_map = {
        "序号": "序号",
        "库名": "库名",
        "表英文名": "表英文名",
        "表中文名": "表中文名",
        "表字段英文名": "表字段英文名",
        "表字段中文名": "表字段中文名",
        "字段类型": "字段类型",
        "长度": "长度",
        "精度": "精度",
        "是否分区字段": "是否分区字段",
        "描述": "描述",
    }
    # 只保留需要的列
    keep = [c for c in col_map if c in df.columns]
    df = df[keep].copy()
    return df

def convert_one_table(df: pd.DataFrame, output_dir: str):
    """把一个表的数据转换成标准 xlsx"""
    if df.empty:
        print("  警告：数据为空，跳过")
        return

    # 取第一行作为表级信息
    first = df.iloc[0]
    table_en = to_lower(first.get("表英文名", ""))
    table_cn = str(first.get("表中文名", "")).strip() if pd.notna(first.get("表中文名")) else ""

    if not table_en:
        print("  错误：找不到表英文名，跳过")
        return

    # ---------- Sheet1: 实体表模板 ----------
    # 主题固定写 “1. 基础能力域”（可根据实际需要调整）
    sheet1_data = [
        ["1. 基础能力域", table_en, table_cn, ""]
    ]

    # ---------- Sheet2: 字段模板 ----------
    sheet2_rows = []
    for _, row in df.iterrows():
        field_en = to_lower(row.get("表字段英文名", ""))
        field_cn = clean_chinese_name(row.get("表字段中文名", ""))
        field_type = normalize_type(row.get("字段类型", "string"))

        # 长度、精度
        length = row.get("长度", "")
        precision = row.get("精度", "0")
        try:
            length = int(float(length)) if pd.notna(length) and str(length).strip() else ""
        except Exception:
            length = ""
        try:
            precision = int(float(precision)) if pd.notna(precision) and str(precision).strip() else 0
        except Exception:
            precision = 0

        is_pk = "否"          # 输入文件中没有主键信息，统一写否
        is_part = is_partition(row.get("是否分区字段", "否"))
        not_null = "否"       # 输入文件中没有非空信息，统一写否
        desc = str(row.get("描述", "")).strip() if pd.notna(row.get("描述")) else ""

        sheet2_rows.append([
            table_en,           # 表名(必填) → 小写
            field_en,           # 实体属性英文名(必填) → 小写
            field_cn,           # 实体属性中文名(必填) → 清洗后
            field_type,         # 实体属性类型(必填)
            length,             # 长度
            precision,          # 精度
            is_pk,              # 是否主键
            is_part,            # 是否分区
            not_null,           # 非空
            desc                # 描述
        ])

    # ---------- 写 xlsx ----------
    wb = Workbook()

    # Sheet1
    ws1 = wb.active
    ws1.title = "实体表模板"
    ws1.append(SHEET1_HEADERS)
    for r in sheet1_data:
        ws1.append(r)

    # Sheet2
    ws2 = wb.create_sheet("字段模板")
    ws2.append(SHEET2_HEADERS)
    for r in sheet2_rows:
        ws2.append(r)

    # 简单样式
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    for ws in (ws1, ws2):
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin
        # 自动列宽（简单估算）
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value)) if cell.value else 0)
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    out_path = Path(output_dir) / f"{table_en}.xlsx"
    wb.save(out_path)
    print(f"  已生成: {out_path}")

# ===================== 主流程 =====================
def process_file(file_path: str, output_dir: str):
    print(f"处理文件: {file_path}")
    try:
        df = read_input_file(file_path)
        # 一个文件可能只包含一张表（按表英文名分组）
        if "表英文名" not in df.columns:
            print("  错误：找不到【表英文名】列")
            return
        for table_name, group in df.groupby("表英文名"):
            convert_one_table(group, output_dir)
    except Exception as e:
        print(f"  处理失败: {e}")

def process_zip(zip_path: str, output_dir: str):
    print(f"处理 zip 包: {zip_path}")
    with tempfile.TemporaryDirectory() as tmpdir:
        with zipfile.ZipFile(zip_path, "r") as zf:
            zf.extractall(tmpdir)
        for root, _, files in os.walk(tmpdir):
            for f in files:
                if f.lower().endswith((".xls", ".xlsx")) and not f.startswith("~$"):
                    process_file(os.path.join(root, f), output_dir)

def main():
    parser = argparse.ArgumentParser(description="模型详情表 → 标准实体表模板转换工具")
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--file", help="单个输入文件路径（.xls 或 .xlsx）")
    group.add_argument("--zip", help="输入 zip 包路径（可包含多个同格式文件）")
    parser.add_argument("--output", required=True, help="输出文件夹路径")
    args = parser.parse_args()

    output_dir = Path(args.output)
    output_dir.mkdir(parents=True, exist_ok=True)

    if args.file:
        process_file(args.file, str(output_dir))
    else:
        process_zip(args.zip, str(output_dir))

    print("全部完成。")

if __name__ == "__main__":
    main()