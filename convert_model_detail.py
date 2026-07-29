#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
模型详情 Excel 转换工具
======================
读取 "模型详情-库：xxx-表：yyy.xls" 格式的 Excel 文件，
根据内容生成以表名命名的 .xlsx 文件，包含 "实体表模板" 和 "字段模板" 两个 sheet。

用法:
    python convert_model_detail.py <输入文件路径> [输出目录]

示例:
    python convert_model_detail.py "C:\\Users\\xxx\\Desktop\\模型详情-库：ap_tenant_user7-表：hall_estimate_d.xls"
    python convert_model_detail.py "C:\\Users\\xxx\\Desktop\\模型详情-库：ap_tenant_user7-表：hall_estimate_d.xls" "D:\\output"

无参数运行时，自动查找桌面上的 模型详情*.xls 文件。

依赖: xlrd, pandas, openpyxl  (pip install xlrd pandas openpyxl)
"""

import os
import re
import sys
import glob

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.colors import Color
from openpyxl.worksheet.datavalidation import DataValidation


# ==================== 配置区 ====================

# 主题默认值（输入文件中不包含此信息，使用固定默认值）
DEFAULT_THEME = "1. 基础能力域"

# 是否主键默认值（输入文件中不包含此信息）
DEFAULT_IS_PK = "否"

# 非空默认值（输入文件中不包含此信息）
DEFAULT_NOT_NULL = "否"

# ---- 表头样式（与模板一致）----
HEADER_FONT = Font(bold=True, size=12)
HEADER_FILL = PatternFill(fill_type="solid", fgColor=Color(indexed=57))

# ---- 实体表模板 sheet ----
ENTITY_SHEET_NAME = "实体表模板"
ENTITY_HEADERS = ["主题(必填)", "实体英文名(必填)", "实体中文名(必填)", "描述"]
ENTITY_COL_WIDTHS = {"A": 30.0}

# ---- 字段模板 sheet ----
FIELD_SHEET_NAME = "字段模板"
FIELD_HEADERS = [
    "表名(必填)",
    "实体属性英文名(必填)",
    "实体属性中文名(必填)",
    "实体属性类型(必填)",
    "实体属性类型长度(整型数字)",
    "精度(整型数字)",
    "是否主键(必填)",
    "是否分区(必填)",
    "非空(必填)",
    "描述",
]
# 实体属性类型下拉列表
FIELD_TYPE_OPTIONS = '"string,bigint,double,decimal,boolean,date,timestamp"'

# 输入文件中的类型 → 标准类型映射（映射到下拉列表中的值）
FIELD_TYPE_MAP = {
    "string": "string",
    "bigint": "bigint",
    "int": "bigint",
    "double": "double",
    "decimal": "decimal",
    "boolean": "boolean",
    "date": "date",
    "timestamp": "timestamp",
}

FIELD_COL_WIDTHS = {
    "A": 30.0,
    "B": 14.89,
    "C": 13.44,
    "D": 19.44,
    "E": 30.0,
    "F": 20.0,
    "J": 30.0,
}

# 输入文件中需要使用的列名
COL_SEQ = "序号"
COL_SCHEMA = "库名"
COL_TABLE_EN = "表英文名"
COL_TABLE_CN = "表中文名"
COL_FIELD_EN = "表字段英文名"
COL_FIELD_CN = "表字段中文名"
COL_FIELD_TYPE = "字段类型"
COL_LENGTH = "长度"
COL_PRECISION = "精度"
COL_IS_PARTITION = "是否分区字段"
COL_DESC = "描述"


# ==================== 工具函数 ====================

def to_clean_str(val) -> str:
    """转为清理后的字符串；NaN / None / 空白 返回空串。"""
    if val is None:
        return ""
    s = str(val).strip()
    if s.lower() in ("nan", "none", ""):
        return ""
    return s


def to_int_str(val) -> str:
    """将数值（含浮点字符串如 '20.0'）转为整数字符串 '20'；异常返回 '0'。"""
    s = to_clean_str(val)
    if not s:
        return "0"
    try:
        return str(int(float(s)))
    except (ValueError, TypeError):
        return "0"


def clean_field_cn(val) -> str:
    """
    清理实体属性中文名，确保符合以下规则：
      1. 任何位置不能包含空格 — 直接删除所有空白字符
      2. 不能以数字或其他非中英文文字开头 — 删除开头非中英文字符
      3. 首字符之后允许中文、英文、数字、下划线、中杠；
         标点符号、括号等其他符号 — 直接删除
    """
    s = to_clean_str(val)
    # 1. 删除所有空白字符（含空格、制表符等）
    s = re.sub(r'\s', '', s)
    # 2. 仅保留中文、英文、数字、下划线、中杠，其余字符（标点、括号等）直接删除
    s = re.sub(r'[^\u4e00-\u9fffa-zA-Z0-9_\-]', '', s)
    # 3. 删除开头的非中英文字符（数字、下划线、中杠不能开头）
    s = re.sub(r'^[^a-zA-Z\u4e00-\u9fff]+', '', s)
    return s


def read_model_detail(input_file: str) -> pd.DataFrame:
    """读取模型详情 Excel 文件，返回 DataFrame（全部为字符串类型）。"""
    ext = os.path.splitext(input_file)[1].lower()
    if ext == ".xls":
        engine = "xlrd"
    elif ext in (".xlsx", ".xlsm"):
        engine = "openpyxl"
    else:
        raise ValueError(f"不支持的文件格式: {ext}，仅支持 .xls / .xlsx")

    df = pd.read_excel(input_file, sheet_name=0, engine=engine, dtype=str)
    df.columns = df.columns.str.strip()
    return df


# ==================== 核心逻辑 ====================

def generate_output(df: pd.DataFrame, output_dir: str) -> list:
    """
    根据模型详情 DataFrame 生成输出 Excel 文件。
    一个输入文件中可能包含多张表，按 (库名, 表英文名, 表中文名) 分组，每张表生成一个文件。
    返回 [(输出路径, 字段数), ...]
    """
    results = []

    # 按序号排序，保证字段顺序与原始一致
    df = df.copy()
    df[COL_SEQ] = pd.to_numeric(df[COL_SEQ], errors="coerce")
    df = df.sort_values(COL_SEQ)

    # 按 (库名, 表英文名, 表中文名) 分组
    for (schema, table_en, table_cn), group in df.groupby(
        [COL_SCHEMA, COL_TABLE_EN, COL_TABLE_CN]
    ):
        # 小写化：库名、表名、字段名
        table_en_lower = to_clean_str(table_en).lower()
        table_cn_str = to_clean_str(table_cn)

        if not table_en_lower:
            print("  警告: 跳过空表名行")
            continue

        output_file = os.path.join(output_dir, f"{table_en_lower}.xlsx")

        # 创建工作簿
        wb = openpyxl.Workbook()

        # -------- Sheet 1: 实体表模板 --------
        ws1 = wb.active
        ws1.title = ENTITY_SHEET_NAME

        for col_idx, header in enumerate(ENTITY_HEADERS, 1):
            cell = ws1.cell(row=1, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL

        ws1.cell(row=2, column=1, value=DEFAULT_THEME)
        ws1.cell(row=2, column=2, value=table_en_lower)
        ws1.cell(row=2, column=3, value=table_cn_str)
        ws1.cell(row=2, column=4, value=None)

        for col_letter, width in ENTITY_COL_WIDTHS.items():
            ws1.column_dimensions[col_letter].width = width

        # -------- Sheet 2: 字段模板 --------
        ws2 = wb.create_sheet(FIELD_SHEET_NAME)

        for col_idx, header in enumerate(FIELD_HEADERS, 1):
            cell = ws2.cell(row=1, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL

        for row_idx, (_, row) in enumerate(group.iterrows(), 2):
            field_en = to_clean_str(row[COL_FIELD_EN]).lower()   # 字段名小写
            field_cn = clean_field_cn(row[COL_FIELD_CN])
            field_type = to_clean_str(row[COL_FIELD_TYPE])
            field_type = FIELD_TYPE_MAP.get(field_type.lower(), field_type.lower())
            length = to_int_str(row[COL_LENGTH])
            precision = to_int_str(row[COL_PRECISION])
            is_partition = to_clean_str(row[COL_IS_PARTITION])
            desc = to_clean_str(row.get(COL_DESC, ""))

            ws2.cell(row=row_idx, column=1, value=table_en_lower)       # 表名(小写)
            ws2.cell(row=row_idx, column=2, value=field_en)             # 实体属性英文名(小写)
            ws2.cell(row=row_idx, column=3, value=field_cn)             # 实体属性中文名
            ws2.cell(row=row_idx, column=4, value=field_type)           # 实体属性类型
            ws2.cell(row=row_idx, column=5, value=length)               # 长度
            ws2.cell(row=row_idx, column=6, value=precision)            # 精度
            ws2.cell(row=row_idx, column=7, value=DEFAULT_IS_PK)        # 是否主键
            ws2.cell(row=row_idx, column=8, value=is_partition)         # 是否分区
            ws2.cell(row=row_idx, column=9, value=DEFAULT_NOT_NULL)     # 非空
            ws2.cell(row=row_idx, column=10, value=desc if desc else None)  # 描述

        for col_letter, width in FIELD_COL_WIDTHS.items():
            ws2.column_dimensions[col_letter].width = width

        # 添加"实体属性类型"列下拉列表验证（D列，从第2行到最后一行）
        field_type_dv = DataValidation(
            type="list",
            formula1=FIELD_TYPE_OPTIONS,
            allow_blank=True,
        )
        field_type_dv.error = "请从下拉列表中选择类型"
        field_type_dv.errorTitle = "输入错误"
        field_type_dv.prompt = "请选择实体属性类型"
        field_type_dv.promptTitle = "实体属性类型"
        last_row = ws2.max_row
        field_type_dv.add(f"D2:D{last_row}")
        ws2.add_data_validation(field_type_dv)

        # 保存
        wb.save(output_file)
        field_count = len(group)
        results.append((output_file, field_count))
        print(f"  已生成: {output_file}  (表: {table_en_lower}, 字段数: {field_count})")

    return results


# ==================== 主入口 ====================

def main():
    args = sys.argv[1:]

    # 无参数时自动查找桌面上的模型详情文件
    if not args:
        desktop = os.path.join(os.path.expanduser("~"), "Desktop")
        candidates = glob.glob(os.path.join(desktop, "模型详情*.xls"))
        if candidates:
            print(f"未指定参数，自动在桌面找到 {len(candidates)} 个文件:")
            for f in candidates:
                print(f"  {f}")
            print()
            input_files = candidates
            output_dir = desktop
        else:
            print("用法: python convert_model_detail.py <输入文件路径> [输出目录]")
            print("未找到输入文件，请指定路径。")
            sys.exit(1)
    else:
        # 判断最后一个参数是否为输出目录
        if len(args) >= 2 and os.path.isdir(args[-1]):
            output_dir = args[-1]
            file_args = args[:-1]
        else:
            output_dir = None
            file_args = args

        # 展开通配符
        input_files = []
        for arg in file_args:
            if "*" in arg or "?" in arg:
                input_files.extend(glob.glob(arg))
            else:
                input_files.append(arg)

        if output_dir is None:
            # 默认输出到第一个输入文件所在目录
            output_dir = os.path.dirname(input_files[0]) or "."

    print(f"输出目录: {output_dir}")
    print(f"待处理文件: {len(input_files)} 个")
    print()

    total = 0
    for input_file in input_files:
        if not os.path.exists(input_file):
            print(f"错误: 文件不存在 - {input_file}")
            continue

        print(f"正在处理: {input_file}")
        try:
            df = read_model_detail(input_file)
            results = generate_output(df, output_dir)
            total += len(results)
        except Exception as e:
            print(f"  错误: {e}")

    print(f"\n完成! 共生成 {total} 个文件。")


if __name__ == "__main__":
    main()
